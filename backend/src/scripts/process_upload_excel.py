import os
import sys
import re
import json
import time
import subprocess
import pandas as pd
import openpyxl
from datetime import datetime

PROJECT_ID = "project-silvia-500416"

def strip_accents(s):
    if not s:
        return ""
    # Map accented vowels to normal ones
    s = str(s)
    accents = {'Á': 'A', 'É': 'E', 'Í': 'I', 'Ó': 'O', 'Ú': 'U', 'Ü': 'U', 'Ñ': 'N'}
    for char, rep in accents.items():
        s = s.replace(char, rep)
    return s

def clean_plate(plate):
    if not plate or not isinstance(plate, str):
        return ""
    cleaned = re.sub(r'[^A-Z0-9]', '', plate.upper())
    if cleaned in ('NA', 'TOTAL', 'RESUMEN', 'PROMEDIO', 'FLETE', 'GUIA', 'GUIDE', 'CANT', 'CANTERA'):
        return ""
    return cleaned

def parse_date(val):
    if val is None or pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val.date()
    if hasattr(val, 'date'):
        return val.date()
    val_str = str(val).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S", "%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%d/%m/%Y %H:%M:%S", "%d/%m/%y", "%d-%m-%Y"):
        try:
            return datetime.strptime(val_str.split()[0], fmt).date()
        except:
            pass
    return None

def try_float(val):
    if val is None or pd.isna(val):
        return 0.0
    try:
        if isinstance(val, str):
            val = val.replace(",", "")
        return float(val)
    except:
        return 0.0

def load_placa_to_id():
    veh_path = os.environ.get("VEHICLES_CSV_PATH", "/home/josue/Documents/SILVIA/processed_data/vehiculos.csv")
    placa_to_id = {}
    if os.path.exists(veh_path):
        df_veh = pd.read_csv(veh_path)
        for row in df_veh.itertuples():
            cp = clean_plate(row.placa)
            if cp:
                placa_to_id[cp] = row.id
    return placa_to_id

def is_valid_factura(val):
    if val is None or pd.isna(val):
        return False
    val_str = str(val).strip().upper()
    if not val_str or val_str in ("FACTURA", "TOTAL", "DEP", "CHOFER", "PRODUCTO", "GALON", "P.U", "IMPORTE", "ORDEN", "NONE"):
        return False
    if "TOTAL" in val_str or "SALDO" in val_str or "RUC" in val_str or "ORDEN" in val_str:
        return False
    return True

def process_diesel(file_path, placa_to_id):
    # Convert if xls to xlsx
    temp_xlsx = None
    if file_path.lower().endswith(".xls"):
        temp_dir = os.path.dirname(file_path)
        sys.stderr.write(f"[INFO] Converting XLS to XLSX...\n")
        subprocess.run([
            "soffice", "--headless",
            "--convert-to", "xlsx",
            "--outdir", temp_dir,
            file_path
        ], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        
        filename = os.path.basename(file_path)
        xlsx_name = filename[:-4] + ".xlsx"
        temp_xlsx = os.path.join(temp_dir, xlsx_name)
        file_path = temp_xlsx
        
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    records = []
    
    # Headers are:
    # Col 1: Flota, Col 2: Cod. Autoriz., Col 3: Sitio, Col 4: ID Contribuyente, Col 5: Num. Trans., Col 6: Fecha, Col 7: Hora, Col 8: SubCuenta (plate), Col 10: Combustible, Col 11: Volumen Despachado (galones), Col 12: Monto Despachado, Col 13: Precio unitario Despachado
    for r in range(2, sheet.max_row + 1):
        fecha_raw = sheet.cell(row=r, column=6).value
        fecha_parsed = parse_date(fecha_raw)
        if not fecha_parsed:
            continue
            
        hora_val = sheet.cell(row=r, column=7).value
        # Format time to string
        if isinstance(hora_val, datetime):
            hora_str = hora_val.strftime("%H:%M:%S")
        elif hasattr(hora_val, 'strftime'):
            hora_str = hora_val.strftime("%H:%M:%S")
        else:
            hora_str = str(hora_val or "00:00:00").strip()
            
        placa_raw = sheet.cell(row=r, column=8).value
        if placa_raw:
            # Placa can be like "AFX735 - AFX735"
            placa_clean = clean_plate(str(placa_raw).split("-")[0])
        else:
            placa_clean = ""
            
        vehiculo_id = placa_to_id.get(placa_clean, placa_clean)
        
        galones = try_float(sheet.cell(row=r, column=11).value)
        monto = try_float(sheet.cell(row=r, column=12).value)
        pu = try_float(sheet.cell(row=r, column=13).value)
        
        # Unique ID for row
        unique_id = f"ds_{fecha_parsed.strftime('%Y%m%d')}_{r}_{placa_clean}"
        
        records.append({
            "id": unique_id,
            "vehiculo_id": vehiculo_id,
            "placa": placa_clean,
            "fecha": fecha_parsed.strftime("%Y-%m-%d"),
            "hora": f"{fecha_parsed.strftime('%Y-%m-%d')} {hora_str}",
            "galones_despachados": round(galones, 3),
            "precio_unitario": round(pu, 3),
            "monto_despachado": round(monto, 2)
        })
        
    wb.close()
    if temp_xlsx and os.path.exists(temp_xlsx):
        try:
            os.unlink(temp_xlsx)
        except:
            pass
            
    return records

def process_gnv(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    # 1. Detect format
    is_formato_a = False
    header_row_idx = None
    for r in range(1, 20):
        row_vals = [str(sheet.cell(row=r, column=c).value).strip().upper() for c in range(1, sheet.max_column + 1)]
        if any("FECHA DESPACHO" in x for x in row_vals) and any("PLACA" in x for x in row_vals):
            is_formato_a = True
            header_row_idx = r
            break
            
    records = []
    
    if is_formato_a:
        sys.stderr.write("[INFO] Parsing GNV Formato A (Estado de cuenta)...\n")
        headers = [str(sheet.cell(row=header_row_idx, column=c).value).strip() for c in range(1, sheet.max_column + 1)]
        col_map = {}
        for i, h in enumerate(headers):
            if h is None or h == 'None':
                continue
            h_clean = strip_accents(h.upper().replace(" ", "").replace("°", "").replace("\n", ""))
            if "FECHADESPACHO" in h_clean:
                col_map["fecha"] = i + 1
            elif "DOCUMENTO" in h_clean:
                col_map["factura"] = i + 1
            elif "PLACA" in h_clean:
                col_map["placa"] = i + 1
            elif "PRECIOM3" in h_clean:
                col_map["precio_unitario"] = i + 1
            elif "CANT.M3" in h_clean or "CANTM3" in h_clean:
                col_map["m3"] = i + 1
            elif h_clean == "TOTAL":
                col_map["monto_total"] = i + 1
                
        for r in range(header_row_idx + 1, sheet.max_row + 1):
            fecha = parse_date(sheet.cell(row=r, column=col_map.get("fecha", 99)).value)
            factura = sheet.cell(row=r, column=col_map.get("factura", 99)).value
            placa = sheet.cell(row=r, column=col_map.get("placa", 99)).value
            m3_val = try_float(sheet.cell(row=r, column=col_map.get("m3", 99)).value)
            pu = try_float(sheet.cell(row=r, column=col_map.get("precio_unitario", 99)).value)
            total = try_float(sheet.cell(row=r, column=col_map.get("monto_total", 99)).value)
            
            if fecha is not None and is_valid_factura(factura):
                placa_clean = clean_plate(placa)
                records.append({
                    "fecha": fecha.strftime("%Y-%m-%d"),
                    "factura": str(factura).strip(),
                    "placa": placa_clean,
                    "m3": round(m3_val, 2),
                    "precio_unitario": round(pu, 4),
                    "monto_total": round(total, 2)
                })
    else:
        sys.stderr.write("[INFO] Parsing GNV Formato B (Reporte / Judy)...\n")
        # Format B has multiple tables, let's scan all rows
        for r in range(1, sheet.max_row + 1):
            # Left Table (Col 1-8)
            fecha_l = parse_date(sheet.cell(row=r, column=1).value)
            factura_l = sheet.cell(row=r, column=2).value
            placa_l = sheet.cell(row=r, column=4).value
            galon_l = try_float(sheet.cell(row=r, column=5).value)
            precio_auth_l = try_float(sheet.cell(row=r, column=8).value)
            
            if fecha_l is not None and is_valid_factura(factura_l):
                pu_l = 0.0
                if galon_l and precio_auth_l:
                    pu_l = round(precio_auth_l / galon_l, 4)
                records.append({
                    "fecha": fecha_l.strftime("%Y-%m-%d"),
                    "factura": str(factura_l).strip(),
                    "placa": clean_plate(placa_l),
                    "m3": round(galon_l, 2),
                    "precio_unitario": pu_l,
                    "monto_total": round(precio_auth_l, 2)
                })
                
            # Right Table (Col 10-17)
            if sheet.max_column >= 17:
                fecha_r = parse_date(sheet.cell(row=r, column=10).value)
                factura_r = sheet.cell(row=r, column=11).value
                placa_r = sheet.cell(row=r, column=13).value
                galon_r = try_float(sheet.cell(row=r, column=14).value)
                precio_auth_r = try_float(sheet.cell(row=r, column=17).value)
                
                if fecha_r is not None and is_valid_factura(factura_r):
                    pu_r = 0.0
                    if galon_r and precio_auth_r:
                        pu_r = round(precio_auth_r / galon_r, 4)
                    records.append({
                        "fecha": fecha_r.strftime("%Y-%m-%d"),
                        "factura": str(factura_r).strip(),
                        "placa": clean_plate(placa_r),
                        "m3": round(galon_r, 2),
                        "precio_unitario": pu_r,
                        "monto_total": round(precio_auth_r, 2)
                    })
                    
    wb.close()
    return records

def process_peajes(file_path, placa_to_id):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    sheet = wb.active
    
    # Detemine if we need to skip lines (header row could be further down or at row 1)
    header_row_idx = 1
    for r in range(1, 25):
        val = str(sheet.cell(row=r, column=1).value).strip().upper()
        if "TIPO SERVICIO" in val or "TIPO_SERVICIO" in val:
            header_row_idx = r
            break
            
    headers = [str(sheet.cell(row=header_row_idx, column=c).value).strip() for c in range(1, sheet.max_column + 1)]
    col_map = {}
    for i, h in enumerate(headers):
        if h is None or h == 'None':
            continue
        h_clean = strip_accents(h.upper().replace(" ", "").replace("_", ""))
        if "TIPOSERVICIO" in h_clean:
            col_map["tipo_servicio"] = i + 1
        elif "REDUSO" in h_clean:
            col_map["red_uso"] = i + 1
        elif "PLACA" in h_clean:
            col_map["placa"] = i + 1
        elif "FECHATRANSITO" in h_clean:
            col_map["fecha_transito"] = i + 1
        elif "PUNTOSERVICIO" in h_clean or "PUNTODESERVICIO" in h_clean:
            col_map["punto_servicio"] = i + 1
        elif "COMPROBANTE" in h_clean:
            col_map["comprobante"] = i + 1
        elif "TOTAL" in h_clean:
            col_map["total_servicio"] = i + 1
        elif "SALDO" in h_clean:
            col_map["saldo_final"] = i + 1
            
    records = []
    
    for r in range(header_row_idx + 1, sheet.max_row + 1):
        tipo_serv = sheet.cell(row=r, column=col_map.get("tipo_servicio", 99)).value
        if tipo_serv is None or str(tipo_serv).strip() == "" or "saldo" in str(tipo_serv).lower():
            continue
            
        fecha_raw = sheet.cell(row=r, column=col_map.get("fecha_transito", 99)).value
        fecha_parsed = parse_date(fecha_raw)
        if not fecha_parsed:
            continue
            
        placa_raw = sheet.cell(row=r, column=col_map.get("placa", 99)).value
        placa_clean = clean_plate(placa_raw)
        # Map plate to vehicle ID if exists in lookup
        veh_id = placa_to_id.get(placa_clean, placa_clean)
        
        red = sheet.cell(row=r, column=col_map.get("red_uso", 99)).value
        punto = sheet.cell(row=r, column=col_map.get("punto_servicio", 99)).value
        comprobante = sheet.cell(row=r, column=col_map.get("comprobante", 99)).value
        total = try_float(sheet.cell(row=r, column=col_map.get("total_servicio", 99)).value)
        saldo = try_float(sheet.cell(row=r, column=col_map.get("saldo_final", 99)).value)
        
        records.append({
            "tipo_servicio": str(tipo_serv).strip(),
            "red_uso": str(red or "").strip(),
            "placa": veh_id, # Must match vehicle ID to maintain BQ relationships
            "placa_raw": placa_clean,
            "fecha_transito": fecha_parsed.strftime("%Y-%m-%d"),
            "punto_servicio": str(punto or "").strip(),
            "comprobante": str(comprobante or "").strip(),
            "total_servicio": round(total, 2),
            "saldo_final": round(saldo, 2)
        })
        
    wb.close()
    return records

def main():
    if len(sys.argv) < 3:
        sys.stderr.write("Usage: python process_upload_excel.py <document_type> <file_path>\n")
        sys.exit(1)
        
    doc_type = sys.argv[1].upper()
    file_path = sys.argv[2]
    
    if not os.path.exists(file_path):
        sys.stderr.write(f"File not found: {file_path}\n")
        sys.exit(1)
        
    placa_to_id = load_placa_to_id()
    
    records = []
    try:
        if doc_type == "DIESEL":
            records = process_diesel(file_path, placa_to_id)
        elif doc_type == "GNV":
            records = process_gnv(file_path)
        elif doc_type == "PEAJES":
            records = process_peajes(file_path, placa_to_id)
        else:
            sys.stderr.write(f"Unknown document type: {doc_type}\n")
            sys.exit(1)
            
        print(json.dumps(records))
    except Exception as e:
        sys.stderr.write(f"Error parsing file: {e}\n")
        sys.exit(1)

if __name__ == "__main__":
    main()
